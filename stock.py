import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.protection import WorkbookProtection
import shutil
from tabulate import tabulate
from datetime import date, datetime
from dateutil.relativedelta import relativedelta
import sys
sys.path.append("/Applications/Wind API.app/Contents/python")
from WindPy import w

w.start()

while True:
	file_route=''	
#	interval=input('Type 1 for yesterday and 0 for today: ')
	stockcode=input('The stock code is: ')
	today = date.today() - relativedelta(days=int(1))
	tradedate=today.strftime('%Y%m%d')
	year=2022
	rptdate=20230930
	marginal_taxrate=0.25
	riskfree_rate_china=0.029
	riskfree_rate_us=0.036	
	enquity_risk_premium_china=0.066
	enquity_risk_premium_us=0.055	
	terminal_risk_premium=0.05
	if stockcode.endswith('sh') or stockcode.endswith('sz') or stockcode.endswith('bj'):
		riskfree_rate=riskfree_rate_china
		terminal_wacc=riskfree_rate+terminal_risk_premium
		###	Mature companies tend to have costs of INVESTED_CAPITAL closer to the market average. 
		### While the riskfree rate + 4.5% is a close approximation of the average, you can use a slightly higher number (riskfree rate + 6%) for mature companies in riskier businesses and a slightly lower number (risfree rate + 4%) for safer companies
		error_code,relative=w.wss(stockcode, ['sec_name', 'report_cur', 'mkt_cap_ard','total_shares', 'close', 'beta_60m','beta_24m', 'pe_ttm', 'pb_mrq', 'pcf_ocf_ttm'], unit=1, tradeDate=tradedate, rptDate=rptdate, year=year, rptType=1,usedf=True)
		error_code,base_1=w.wsd(stockcode, ['tot_oper_rev','ebit2','taxtoebt','fin_int_exp','da_perid','interestdebt','networkingcapital','tot_equity','minority_int','monetary_cap','tradable_fin_assets','fin_assets_chg_compreh_inc','fin_assets_amortizedcost',
			'debt_invest','long_term_eqy_invest','hfs_assets','loans_to_oth_banks','loans_and_adv_granted','net_cash_flows_oper_act', 'cash_pay_acq_const_fiolta',
			'fund_restricted', 'acctandnotes_rcv', 'prepay', 'oth_rcv_tot', 'inventories', 'cont_assets', 'oth_cur_assets', 'acctandnotes_payable', 'adv_from_cust', 'cont_liab', 'oth_payable_tot', 'acc_exp', 'empl_ben_payable'], 'ED-5Y', '2022-12-31', unit=1, rptType=1, Period='Y', Days='ALLDAYS', usedf=True)
		error_code,driver_1=w.wsd(stockcode, ['yoy_tr','taxtoebt','roe_avg', 'dividendyield2', 'debttoassets'], 'ED-5Y', '2022-12-31', unit=1, rptType=1, Period='Y', Days='ALLDAYS', usedf=True)
		error_code,base_2=w.wsd(stockcode, ['tot_oper_rev','ebit2','taxtoebt','fin_int_exp','da_perid','interestdebt','networkingcapital','tot_equity','minority_int','monetary_cap','tradable_fin_assets','fin_assets_chg_compreh_inc','fin_assets_amortizedcost',
			'debt_invest','long_term_eqy_invest','hfs_assets','loans_to_oth_banks','loans_and_adv_granted','net_cash_flows_oper_act', 'cash_pay_acq_const_fiolta',
			'fund_restricted', 'acctandnotes_rcv', 'prepay', 'oth_rcv_tot', 'inventories', 'cont_assets', 'oth_cur_assets', 'acctandnotes_payable', 'adv_from_cust', 'cont_liab', 'oth_payable_tot', 'acc_exp', 'empl_ben_payable'], 'ED-2Q', '2023-09-30', unit=1, rptType=1, Period='Q', Days='ALLDAYS', usedf=True)
		error_code,driver_2=w.wsd(stockcode, ['yoy_tr','taxtoebt','roe_avg', 'dividendyield2', 'debttoassets'], 'ED-2Q', '2023-09-30', unit=1, rptType=1, Period='Q', Days='ALLDAYS', usedf=True)
		num_observations = len(base_2)
		date_range = pd.date_range(start='2023-03-31', periods=num_observations, freq='3M').strftime('%Y-%m-%d')
		base_2.set_index(date_range, inplace=True)
		driver_2.set_index(date_range, inplace=True)	
		base=pd.concat([base_1, base_2])
		driver=pd.concat([driver_1, driver_2])
		base['REVENUE']=base['TOT_OPER_REV']
		base['REVENUE_CHANGE']=base['REVENUE'].rolling(window=2).apply(lambda x: x.iloc[1] - x.iloc[0])			
		base['EBIT']=base['EBIT2']		
		base['INTEREST']=base['FIN_INT_EXP']
		base['DA']=base['DA_PERID']					
		base['CASH']=base['MONETARY_CAP']
		base['NETWC']=base[['ACCTANDNOTES_RCV', 'PREPAY', 'OTH_RCV_TOT', 'INVENTORIES', 'CONT_ASSETS', 'OTH_CUR_ASSETS']].sum(axis=1)-base[['ACCTANDNOTES_PAYABLE', 'ADV_FROM_CUST', 'CONT_LIAB', 'OTH_PAYABLE_TOT', 'ACC_EXP','EMPL_BEN_PAYABLE']].sum(axis=1)
		base['NETWC_CHANGE']=base['NETWC'].rolling(window=2).apply(lambda x: x.iloc[1] - x.iloc[0])				
		base['INVESTMENTS']=base[['TRADABLE_FIN_ASSETS','FIN_ASSETS_CHG_COMPREH_INC','FIN_ASSETS_AMORTIZEDCOST','DEBT_INVEST','LONG_TERM_EQY_INVEST','HFS_ASSETS','LOANS_TO_OTH_BANKS','LOANS_AND_ADV_GRANTED']].sum(axis=1)
		base['DEBT']=base['INTERESTDEBT']
		base['EQUITY']=base['TOT_EQUITY']
		base['MINORITY']=base['MINORITY_INT']
		base['AVGDEBT']=base['DEBT'].rolling(2).mean()
		base['INVESTED_CAPITAL']=base[['EQUITY','DEBT']].sum(axis=1)-base['CASH']-base['INVESTMENTS']
		base['IC_CHANGE']=base['INVESTED_CAPITAL'].rolling(window=2).apply(lambda x: x.iloc[1] - x.iloc[0])	
		base['OPER_CF']=base['NET_CASH_FLOWS_OPER_ACT']
		base['CAPEX']=base['CASH_PAY_ACQ_CONST_FIOLTA']	
		base['FCFF']=base['OPER_CF']-base['CAPEX']							
		base['REINVESTMENT']=base['CAPEX']-base['DA']+base['NETWC_CHANGE']	
#		driver['EV/FCFF']=relative['MKT_CAP_ARD'].tolist()[0]/base['FCFF']	
		driver['EBIT_MARGIN']=base['EBIT']/base['REVENUE']*100
		driver['DEBTCOST']=base['INTEREST']/base['AVGDEBT']*100
		driver['ROIC']=(base['EBIT']*(100-base['TAXTOEBT'])/100)/base['INVESTED_CAPITAL']*100	
		base['ROIC']=driver['ROIC']
		driver['TAXTOEBT']=base['TAXTOEBT']
		effective_taxrate=driver['TAXTOEBT'].mean()/100		
		driver['REVENUE/INVESTED_CAPITAL']=base['REVENUE']/base['INVESTED_CAPITAL']	
		driver['△REVENUE/△INVESTED_CAPITAL']=base['REVENUE_CHANGE']/base['IC_CHANGE']
		driver['REINVESTMENT_RATE']=base['REINVESTMENT']/(base['EBIT']*(1-effective_taxrate))*100
		driver['EQUITY_W']=relative['MKT_CAP_ARD'].tolist()[0]/(base['DEBT']+relative['MKT_CAP_ARD'].tolist()[0])
		driver['DEBT_W']=base['DEBT']/(base['DEBT']+relative['MKT_CAP_ARD'].tolist()[0])
		if relative['BETA_60M'].tolist()[0]!=None:
			beta=relative['BETA_60M'].tolist()[0]
		elif relative['BETA_24M'].tolist()[0]!=None:
			beta=relative['BETA_24M'].tolist()[0]
		else:
			beta=1		
		driver['WACC']=((riskfree_rate+enquity_risk_premium_china*beta)*driver['EQUITY_W']+driver['DEBTCOST']/100*(1-marginal_taxrate)*driver['DEBT_W'])*100

	elif stockcode.endswith('hk'):
		riskfree_rate=riskfree_rate_china
		terminal_wacc=riskfree_rate+terminal_risk_premium
		###	Mature companies tend to have costs of INVESTED_CAPITAL closer to the market average. 
		### While the riskfree rate + 4.5% is a close approximation of the average, you can use a slightly higher number (riskfree rate + 6%) for mature companies in riskier businesses and a slightly lower number (risfree rate + 4%) for safer companies
		error_code,relative=w.wss(stockcode, ['sec_name', 'report_cur', 'mkt_cap_ard','total_shares', 'close', 'beta_60m','beta_24m', 'pe_ttm', 'pb_mrq', 'pcf_ocf_ttm'], unit=1, tradeDate=tradedate, rptDate=rptdate, year=year, rptType=1,usedf=True)
		error_code,base_1=w.wsd(stockcode, ['wgsd_sales_oper','ebit2','taxtoebt','wgsd_int_exp','wgsd_dep_exp_cf','wgsd_interestdebt2','networkingcapital','wgsd_stkhldrs_eq','wgsd_min_int','wgsd_cce','wgsd_invest_trading',
			'wgsd_invest_st_oth','wgsd_invest_htm','wgsd_invest_afs','wgsd_invest_eq', 'wgsd_invest_lt_oth', 'wgsd_oper_cf','wgsd_capex_ff',
			'wgsd_fund_restricted','wgsd_receiv_tot','wgsd_inventories','wgsd_assets_curr_oth','wgsd_pay_acct','wgsd_payment_unearned','wgsd_liabs_curr_oth'], 'ED-5Y', '2022-12-31', "unit=1;rptType=1;currencyType=;Period=Y;Days=ALLDAYS", usedf=True)
		error_code,driver_1=w.wsd(stockcode, ['yoy_tr','taxtoebt','roe_avg', 'dividendyield2', 'debttoassets'], 'ED-5Y', '2022-12-31', "unit=1;rptType=1;currencyType=;Period=Y;Days=ALLDAYS", usedf=True)
		error_code,base_2=w.wsd(stockcode, ['wgsd_sales_oper','ebit2','taxtoebt','wgsd_int_exp','wgsd_dep_exp_cf','wgsd_interestdebt2','networkingcapital','wgsd_stkhldrs_eq','wgsd_min_int','wgsd_cce','wgsd_invest_trading',
			'wgsd_invest_st_oth','wgsd_invest_htm','wgsd_invest_afs','wgsd_invest_eq', 'wgsd_invest_lt_oth', 'wgsd_oper_cf','wgsd_capex_ff',
			'wgsd_fund_restricted','wgsd_receiv_tot','wgsd_inventories','wgsd_assets_curr_oth','wgsd_pay_acct','wgsd_payment_unearned','wgsd_liabs_curr_oth'], 'ED-2Q', '2023-09-30', "unit=1;rptType=1;currencyType=;Period=Q;Days=ALLDAYS", usedf=True)
		error_code,driver_2=w.wsd(stockcode, ['yoy_tr','taxtoebt','roe_avg', 'dividendyield2', 'debttoassets'], 'ED-2Q', '2023-09-30', "unit=1;rptType=1;currencyType=;Period=Q;Days=ALLDAYS", usedf=True)
		num_observations = len(base_2)
		date_range = pd.date_range(start='2023-03-31', periods=num_observations, freq='3M').strftime('%Y-%m-%d')
		base_2.set_index(date_range, inplace=True)
		driver_2.set_index(date_range, inplace=True)	
		base=pd.concat([base_1, base_2])
		driver=pd.concat([driver_1, driver_2])
		base['REVENUE']=base['WGSD_SALES_OPER']
		base['REVENUE_CHANGE']=base['REVENUE'].rolling(window=2).apply(lambda x: x.iloc[1] - x.iloc[0])			
		base['EBIT']=base['EBIT2']
		base['INTEREST']=base['WGSD_INT_EXP']
		base['DA']=base['WGSD_DEP_EXP_CF']									
		base['CASH']=base['WGSD_CCE']
		base['NETWC']=base[['WGSD_RECEIV_TOT','WGSD_INVENTORIES','WGSD_ASSETS_CURR_OTH']].sum(axis=1)-base[['WGSD_PAY_ACCT','WGSD_PAYMENT_UNEARNED','WGSD_LIABS_CURR_OTH']].sum(axis=1)
		base['NETWC_CHANGE']=base['NETWC'].rolling(window=2).apply(lambda x: x.iloc[1] - x.iloc[0])					
		base['INVESTMENTS']=base[['WGSD_INVEST_TRADING', 'WGSD_INVEST_ST_OTH','WGSD_INVEST_HTM','WGSD_INVEST_AFS','WGSD_INVEST_EQ', 'WGSD_INVEST_LT_OTH','WGSD_ASSETS_CURR_OTH']].sum(axis=1)
		base['DEBT']=base['WGSD_INTERESTDEBT2']		
		base['EQUITY']=base['WGSD_STKHLDRS_EQ']
		base['MINORITY']=base['WGSD_MIN_INT']
		base['AVGDEBT']=base['DEBT'].rolling(2).mean()
		base['INVESTED_CAPITAL']=base[['EQUITY','DEBT']].sum(axis=1)-base['CASH']-base['INVESTMENTS']
		base['IC_CHANGE']=base['INVESTED_CAPITAL'].rolling(window=2).apply(lambda x: x.iloc[1] - x.iloc[0])
		base['OPER_CF']=base['WGSD_OPER_CF']
		base['CAPEX']=base['WGSD_CAPEX_FF']	
		base['FCFF']=base['OPER_CF']-base['CAPEX']
		base['REINVESTMENT']=base['CAPEX']-base['DA']+base['NETWC_CHANGE']
#		driver['EV/FCFF']=relative['MKT_CAP_ARD'].tolist()[0]/base['FCFF']				
		driver['DEBTCOST']=base['INTEREST']/base['AVGDEBT']*100
		driver['EBIT_MARGIN']=base['EBIT']/base['REVENUE']*100
		driver['ROIC']=(base['EBIT']*(100-base['TAXTOEBT'])/100)/base['INVESTED_CAPITAL']*100
		base['ROIC']=driver['ROIC']
		driver['TAXTOEBT']=base['TAXTOEBT']
		effective_taxrate=driver['TAXTOEBT'].mean()/100		
		driver['REVENUE/INVESTED_CAPITAL']=base['REVENUE']/base['INVESTED_CAPITAL']			
		driver['△REVENUE/△INVESTED_CAPITAL']=base['REVENUE_CHANGE']/base['IC_CHANGE']
		driver['REINVESTMENT_RATE']=base['REINVESTMENT']/(base['EBIT']*(1-effective_taxrate))*100
		driver['EQUITY_W']=relative['MKT_CAP_ARD'].tolist()[0]/(base['DEBT']+relative['MKT_CAP_ARD'].tolist()[0])
		driver['DEBT_W']=base['DEBT']/(base['DEBT']+relative['MKT_CAP_ARD'].tolist()[0])
		if relative['BETA_60M'].tolist()[0]!=None:
			beta=relative['BETA_60M'].tolist()[0]
		elif relative['BETA_24M'].tolist()[0]!=None:
			beta=relative['BETA_24M'].tolist()[0]
		else:
			beta=1			
		driver['WACC']=((riskfree_rate+enquity_risk_premium_china*beta)*driver['EQUITY_W']+driver['DEBTCOST']/100*(1-marginal_taxrate)*driver['DEBT_W'])*100
			
	else:
		riskfree_rate=riskfree_rate_us
		terminal_wacc=riskfree_rate+terminal_risk_premium
		###	Mature companies tend to have costs of INVESTED_CAPITAL closer to the market average. 
		### While the riskfree rate + 4.5% is a close approximation of the average, you can use a slightly higher number (riskfree rate + 6%) for mature companies in riskier businesses and a slightly lower number (risfree rate + 4%) for safer companies
		error_code,relative=w.wss(stockcode, ['sec_name', 'report_cur', 'mkt_cap_ard','total_shares', 'close', 'beta_60m','beta_24m', 'pe_ttm', 'pb_mrq', 'pcf_ocf_ttm'], unit=1, tradeDate=tradedate, rptDate=rptdate, year=year, rptType=1,usedf=True)
		error_code,base_1=w.wsd(stockcode, ['wgsd_sales_oper','ebit2','taxtoebt','wgsd_int_exp','wgsd_dep_exp_cf','wgsd_interestdebt2','wgsd_networkingcapital2','wgsd_stkhldrs_eq','wgsd_min_int','wgsd_cce','wgsd_invest_trading',
			'wgsd_invest_st_oth','wgsd_invest_htm','wgsd_invest_afs','wgsd_invest_eq', 'wgsd_invest_lt_oth', 'wgsd_oper_cf','wgsd_capex_ff',
			'wgsd_fund_restricted','wgsd_receiv_tot','wgsd_inventories','wgsd_assets_curr_oth','wgsd_pay_acct','wgsd_payment_unearned','wgsd_liabs_curr_oth'], 'ED-5Y', '2022-12-31', "unit=1;rptType=1;currencyType=;Period=Y;Days=ALLDAYS", usedf=True)
		error_code,driver_1=w.wsd(stockcode, ['yoy_tr','taxtoebt','roe_avg', 'dividendyield2', 'debttoassets'], 'ED-5Y', '2022-12-31', "unit=1;rptType=1;currencyType=;Period=Y;Days=ALLDAYS", usedf=True)
		error_code,base_2=w.wsd(stockcode, ['wgsd_sales_oper','ebit2','taxtoebt','wgsd_int_exp','wgsd_dep_exp_cf','wgsd_interestdebt2','wgsd_networkingcapital2','wgsd_stkhldrs_eq','wgsd_min_int','wgsd_cce','wgsd_invest_trading',
			'wgsd_invest_st_oth','wgsd_invest_htm','wgsd_invest_afs','wgsd_invest_eq', 'wgsd_invest_lt_oth', 'wgsd_oper_cf','wgsd_capex_ff',
			'wgsd_fund_restricted','wgsd_receiv_tot','wgsd_inventories','wgsd_assets_curr_oth','wgsd_pay_acct','wgsd_payment_unearned','wgsd_liabs_curr_oth'], 'ED-2Q', '2023-09-30', "unit=1;rptType=1;currencyType=;Period=Q;Days=ALLDAYS", usedf=True)
		error_code,driver_2=w.wsd(stockcode, ['yoy_tr','taxtoebt','roe_avg', 'dividendyield2', 'debttoassets'], 'ED-2Q', '2023-09-30', "unit=1;rptType=1;currencyType=;Period=Q;Days=ALLDAYS", usedf=True)
		num_observations = len(base_2)
		date_range = pd.date_range(start='2023-03-31', periods=num_observations, freq='3M').strftime('%Y-%m-%d')
		base_2.set_index(date_range, inplace=True)
		driver_2.set_index(date_range, inplace=True)		
		base=pd.concat([base_1, base_2])
		driver=pd.concat([driver_1, driver_2])
		base['REVENUE']=base['WGSD_SALES_OPER']
		base['REVENUE_CHANGE']=base['REVENUE'].rolling(window=2).apply(lambda x: x.iloc[1] - x.iloc[0])			
		base['EBIT']=base['EBIT2']
		base['INTEREST']=base['WGSD_INT_EXP']
		base['DA']=base['WGSD_DEP_EXP_CF']								
		base['CASH']=base['WGSD_CCE']
		base['NETWC']=base[['WGSD_RECEIV_TOT','WGSD_INVENTORIES','WGSD_ASSETS_CURR_OTH']].sum(axis=1)-base[['WGSD_PAY_ACCT','WGSD_PAYMENT_UNEARNED','WGSD_LIABS_CURR_OTH']].sum(axis=1)
		base['NETWC_CHANGE']=base['NETWC'].rolling(window=2).apply(lambda x: x.iloc[1] - x.iloc[0])					
		base['INVESTMENTS']=base[['WGSD_INVEST_TRADING', 'WGSD_INVEST_ST_OTH','WGSD_INVEST_HTM','WGSD_INVEST_AFS','WGSD_INVEST_EQ', 'WGSD_INVEST_LT_OTH','WGSD_ASSETS_CURR_OTH']].sum(axis=1)
		base['DEBT']=base['WGSD_INTERESTDEBT2']		
		base['EQUITY']=base['WGSD_STKHLDRS_EQ']
		base['MINORITY']=base['WGSD_MIN_INT']
		base['AVGDEBT']=base['DEBT'].rolling(2).mean()
		base['INVESTED_CAPITAL']=base[['EQUITY','DEBT']].sum(axis=1)-base['CASH']-base['INVESTMENTS']
		base['IC_CHANGE']=base['INVESTED_CAPITAL'].rolling(window=2).apply(lambda x: x.iloc[1] - x.iloc[0])
		base['OPER_CF']=base['WGSD_OPER_CF']
		base['CAPEX']=base['WGSD_CAPEX_FF']	
		base['FCFF']=base['OPER_CF']-base['CAPEX']
		base['REINVESTMENT']=base['CAPEX']-base['DA']+base['NETWC_CHANGE']	
#		driver['EV/FCFF']=relative['MKT_CAP_ARD'].tolist()[0]/base['FCFF']				
		driver['DEBTCOST']=base['INTEREST']/base['AVGDEBT']*100
		driver['EBIT_MARGIN']=base['EBIT']/base['REVENUE']*100
		driver['ROIC']=(base['EBIT']*(100-base['TAXTOEBT'])/100)/base['INVESTED_CAPITAL']*100
		base['ROIC']=driver['ROIC']
		driver['TAXTOEBT']=base['TAXTOEBT']
		effective_taxrate=driver['TAXTOEBT'].mean()/100		
		driver['REVENUE/INVESTED_CAPITAL']=base['REVENUE']/base['INVESTED_CAPITAL']
		driver['△REVENUE/△INVESTED_CAPITAL']=base['REVENUE_CHANGE']/base['IC_CHANGE']
		driver['REINVESTMENT_RATE']=base['REINVESTMENT']/(base['EBIT']*(1-effective_taxrate))*100
		driver['EQUITY_W']=relative['MKT_CAP_ARD'].tolist()[0]/(base['DEBT']+relative['MKT_CAP_ARD'].tolist()[0])
		driver['DEBT_W']=base['DEBT']/(base['DEBT']+relative['MKT_CAP_ARD'].tolist()[0])
		if relative['BETA_60M'].tolist()[0]!=None:
			beta=relative['BETA_60M'].tolist()[0]
		elif relative['BETA_24M'].tolist()[0]!=None:
			beta=relative['BETA_24M'].tolist()[0]
		else:
			beta=1			
		driver['WACC']=((riskfree_rate+enquity_risk_premium_us*beta)*driver['EQUITY_W']+driver['DEBTCOST']/100*(1-marginal_taxrate)*driver['DEBT_W'])*100
	
	base_copy=base[['REVENUE','EBIT','CASH','INVESTMENTS','NETWC','DEBT','EQUITY','MINORITY','INVESTED_CAPITAL','CAPEX','DA','NETWC_CHANGE','REINVESTMENT']]/1000000
	base_copy=base_copy.T
	driver_copy=driver[['YOY_TR','EBIT_MARGIN','ROIC','ROE_AVG','DEBTTOASSETS','DIVIDENDYIELD2','REVENUE/INVESTED_CAPITAL','△REVENUE/△INVESTED_CAPITAL','WACC','DEBTCOST','TAXTOEBT']]
	driver_copy=driver_copy.T
	relative_copy=relative.applymap(lambda x: '{:,.2f}'.format(x) if isinstance(x, (int, float)) else x)
#	base_copy=base_copy.applymap(lambda x: '{:,.0f}'.format(x) if isinstance(x, (int, float)) else x)
#	driver_copy=driver_copy.applymap(lambda x: '{:,.1f}'.format(x) if isinstance(x, (int, float)) else x)	

	print(tabulate(relative_copy, headers='keys', tablefmt='psql', numalign='right'))	
	print(tabulate(base_copy, headers='keys', tablefmt='psql', stralign='left', numalign='right',floatfmt=".2f"))
	print(tabulate(driver_copy, headers='keys', tablefmt='psql', stralign='left', numalign='right',floatfmt=".2f"))

	cont = input('Do you want to kick off the company valuation? (y/n): ')
	if cont == 'n':
		exit_program = input('Do you want to exit the program? (y/n): ')
		if exit_program == 'y':
			break
		else:
			continue	
	base_year=int(input('Base year for valuation: '))
	revenue_growth_1=float(input('Compound annual revenue growth rate (next year): '))
	revenue_growth_2=float(input('Compound annual revenue growth rate (FY2-5): '))		
	ebit_margin=float(input('Year 10 target EBIT margin: '))
	convergence=float(input('Years of convergence for target margin: '))
	revenue_INVESTED_CAPITAL_ratio_1=float(input('Revenue to invested capital ratio (next 2 years): '))
	revenue_INVESTED_CAPITAL_ratio_2=float(input('Revenue to invested capital ratio (FY3-5): '))
	revenue_INVESTED_CAPITAL_ratio_3=float(input('Revenue to invested capital ratio (FY5-10): '))
	wacc=float(input('WACC: '))
	cont = input('Do you expect ROIC equal to cost of capital (WACC) beyond year 10? (y/n): ')
	if cont == 'y':
		ronic = terminal_wacc
	else:
		ronic = terminal_wacc + 0.05
### The expected rate of return on new invested INVESTED_CAPITAL (RONIC) should be consistent with expected competitive conditions beyond the explicit forecast period.
### Economic theory suggests that competition will eventually eliminate abnormal returns, so for companies in competitive industries, set RONIC equal to WACC.
### There are some firms with sustainable competitive advantages (brand name, for instance), where the excess returns may continue beyond year 10. 
### If your firm is one of those, you can enter a return on INVESTED_CAPITAL higher than your cost of capital. At the maximum, the excess return should not exceed 5% for a mature firm		
	date_1=datetime(base_year,12,31).date()
	db1=base.loc[date_1]
#	date_2=datetime(2023,9,30).date().strftime('%Y-%m-%d')
#	db2=base.loc[date_2]
	template=file_route + '/' + 'stock valuation template.xlsx'
	filename=file_route + '/' + relative['SEC_NAME'].tolist()[0]+'估值-'+date.today().strftime('%Y%m%d')+'.xlsx'
	shutil.copy(template, filename)
	wb = load_workbook(filename=filename)
	ws1 = wb['Assumptions and Whatif']
	ws2 = wb['WACC model']
	ws3 = wb['Historical data']
	for r in dataframe_to_rows(base_copy, index=True, header=True):
	    ws3.append(r)  
	ws3.insert_rows(ws3.max_row + 1)    	    
	for r in dataframe_to_rows(driver_copy, index=True, header=False):
	    ws3.append(r)	    
	for x in range(1,1):
		for y in range(1,11):
			ws3.cell(row=x, column=y).number_format = 'yyyy-mm-d'
	for x in range(2,30):
		for y in range(1,11):
			ws3.cell(row=x, column=y).number_format = '#,##0.00'
	for z in ['A']:
		ws3.column_dimensions[z].width = 25
	for zz in ['B','C','D','E','F','G','H','I', 'J','K']:
		ws3.column_dimensions[zz].width = 15	
	ws3.cell(row=1,column=1).value='In millions'					
	ws1.cell(row=2,column=3).value=base_year
	ws1.cell(row=3,column=3).value=revenue_growth_1/100
	ws1.cell(row=4,column=3).value=revenue_growth_2/100		
	ws1.cell(row=5,column=3).value=riskfree_rate	
	ws1.cell(row=6,column=3).value=ebit_margin/100	
	ws1.cell(row=7,column=3).value=convergence
	ws1.cell(row=8,column=3).value=revenue_INVESTED_CAPITAL_ratio_1
	ws1.cell(row=9,column=3).value=revenue_INVESTED_CAPITAL_ratio_2
	ws1.cell(row=10,column=3).value=revenue_INVESTED_CAPITAL_ratio_3
	ws1.cell(row=13,column=3).value=ronic	
	ws1.cell(row=14,column=3).value=effective_taxrate	
	ws1.cell(row=11,column=3).value=wacc/100
	ws1.cell(row=12,column=3).value=terminal_wacc
	ws1.cell(row=8,column=6).value=revenue_growth_2/100
	ws1.cell(row=2,column=12).value=ebit_margin/100			
	ws2.cell(row=1,column=1).value=relative['SEC_NAME'].tolist()[0]	
### For the income statement forecast, use last year results.
	ws2.cell(row=4,column=2).value=db1['REVENUE']	
	ws2.cell(row=6,column=2).value=db1['EBIT']
### for the non operating balance sheet items, use the most recent quarterly results.
	ws2.cell(row=22,column=2).value=db1['CASH']
	ws2.cell(row=23,column=2).value=db1['INVESTMENTS']
	ws2.cell(row=25,column=2).value=db1['DEBT']
	ws2.cell(row=26,column=2).value=db1['MINORITY']			
	ws2.cell(row=28,column=2).value=relative['TOTAL_SHARES'].tolist()[0]				
### for invested INVESTED_CAPITAL, use last year end balance since this is compared with after-tax EBIT to calculate ROIC.
	ws2.cell(row=33,column=2).value=db1['INVESTED_CAPITAL']

	wb.active = 0

#	ws.protection.sheet = True
#	ws.protection.password = ''	
#	wb.security = WorkbookProtection()
#	wb.security.workbookPassword = ''
#	wb.security.lockStructure = True
	wb.save(filename)

#	wb = load_workbook(filename=filename)
#	with pd.ExcelWriter(filename, engine='openpyxl', if_sheet_exists='replace', mode='a') as writer:
#		driver_t.to_excel(writer, sheet_name='Historical data', index=False)
#	wb.save(filename)

	cont = input('The valuation process is completed. Do you want to exit the program? (y/n): ')
	if cont == 'y':
		break	